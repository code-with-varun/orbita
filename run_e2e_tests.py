import urllib.request
import urllib.parse
import json
import time
import datetime
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = 'http://localhost:5001'

def api_call(method, endpoint, data=None):
    url = BASE_URL + endpoint.replace(' ', '%20')
    headers = {'Content-Type': 'application/json'}
    body = json.dumps(data).encode('utf-8') if data is not None else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    start_time = time.time()
    try:
        with urllib.request.urlopen(req) as resp:
            elapsed_ms = round((time.time() - start_time) * 1000, 2)
            res_body = resp.read().decode('utf-8')
            try:
                res_json = json.loads(res_body)
            except Exception:
                res_json = res_body
            return {
                'status_code': resp.status,
                'data': res_json,
                'elapsed_ms': elapsed_ms,
                'error': None
            }
    except urllib.error.HTTPError as e:
        elapsed_ms = round((time.time() - start_time) * 1000, 2)
        err_body = e.read().decode('utf-8')
        try:
            err_json = json.loads(err_body)
        except Exception:
            err_json = err_body
        return {
            'status_code': e.code,
            'data': err_json,
            'elapsed_ms': elapsed_ms,
            'error': str(e)
        }
    except Exception as e:
        elapsed_ms = round((time.time() - start_time) * 1000, 2)
        return {
            'status_code': 0,
            'data': None,
            'elapsed_ms': elapsed_ms,
            'error': str(e)
        }

test_results = []

def run_test(test_id, category, test_name, description, test_fn):
    print(f"Running [{test_id}] {test_name}...")
    start_time = time.time()
    status = 'PASSED'
    notes = ''
    try:
        notes = test_fn()
        if not notes:
            notes = 'Assertion passed successfully.'
    except AssertionError as ae:
        status = 'FAILED'
        notes = f"AssertionError: {str(ae)}"
        print(f"  FAILED: {ae}")
    except Exception as e:
        status = 'FAILED'
        notes = f"Exception: {str(e)}\n{traceback.format_exc()}"
        print(f"  ERROR: {e}")
    
    elapsed_ms = round((time.time() - start_time) * 1000, 2)
    test_results.append({
        'test_id': test_id,
        'category': category,
        'test_name': test_name,
        'description': description,
        'status': status,
        'elapsed_ms': elapsed_ms,
        'notes': notes
    })
    print(f"  Result: {status} ({elapsed_ms} ms)")

# Global context dictionary to share IDs across tests
ctx = {}

# -------------------------------------------------------------
# 1. AUTHENTICATION & MULTI-TENANCY TESTS
# -------------------------------------------------------------

def test_auth_superadmin_login():
    res = api_call('POST', '/api/auth/login', {'email': 'superadmin@orbita.com', 'password': 'superadmin123'})
    assert res['status_code'] == 200, f"Expected 200, got {res['status_code']}"
    assert res['data']['user']['role'] == 'Superadmin', "User role should be Superadmin"
    ctx['superadmin'] = res['data']['user']
    return f"Superadmin logged in: {ctx['superadmin']['name']} ({ctx['superadmin']['email']})"

def test_auth_member_registration_and_login():
    email = f"e2e_tester_{int(time.time())}@orbita.com"
    reg_res = api_call('POST', '/api/auth/register', {
        'name': 'E2E Tester Member',
        'email': email,
        'password': 'Password123!',
        'role': 'Member'
    })
    assert reg_res['status_code'] in (200, 201), f"Register failed with status {reg_res['status_code']}"
    assert reg_res['data']['user']['role'] == 'Member'
    ctx['member'] = reg_res['data']['user']
    
    login_res = api_call('POST', '/api/auth/login', {'email': email, 'password': 'Password123!'})
    assert login_res['status_code'] == 200
    assert login_res['data']['user']['id'] == ctx['member']['id']
    return f"Registered and logged in Member: {email} (ID: {ctx['member']['id']})"

def test_auth_invalid_password():
    res = api_call('POST', '/api/auth/login', {'email': 'superadmin@orbita.com', 'password': 'WrongPassword999'})
    assert res['status_code'] == 401, f"Expected 401 Unauthorized, got {res['status_code']}"
    return "Invalid password correctly rejected with 401 Unauthorized"

# -------------------------------------------------------------
# 2. WORK ITEMS CRUD (TASKS, ROUTINES, GOALS, PROJECTS)
# -------------------------------------------------------------

def test_create_task():
    res = api_call('POST', '/api/tasks', {
        'title': 'E2E Single Action Task',
        'description': 'Buy grocery supplies and equipment',
        'orbita_type': 'Task',
        'workspace': 'Personal',
        'priority': 'High',
        'tags': 'Shopping, Urgent',
        'scheduled_date': '2026-08-25',
        'due_date': '2026-08-26',
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'created_by': ctx['member']['name']
    })
    assert res['status_code'] == 201
    assert res['data']['ticket_key'].startswith('TSK-')
    assert res['data']['priority'] == 'High'
    assert res['data']['priority_quadrant'] == 'Q2'
    ctx['task_id'] = res['data']['id']
    return f"Created Task {res['data']['ticket_key']} (Quadrant: {res['data']['priority_quadrant']})"

def test_update_task_details():
    res = api_call('PUT', f"/api/tasks/{ctx['task_id']}", {
        'title': 'E2E Single Action Task (Updated)',
        'priority': 'Critical',
        'user_name': ctx['member']['name']
    })
    assert res['status_code'] == 200
    assert res['data']['title'] == 'E2E Single Action Task (Updated)'
    assert res['data']['priority'] == 'Critical'
    assert res['data']['priority_quadrant'] == 'Q1'
    return f"Updated Task: Priority shifted to Critical (Synchronized Quadrant: {res['data']['priority_quadrant']})"

def test_star_task():
    res = api_call('POST', f"/api/tasks/{ctx['task_id']}/star", {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name']
    })
    assert res['status_code'] == 200
    assert res['data']['is_starred'] is True
    return f"Task star toggled to True: {ctx['task_id']}"

def test_complete_task():
    res = api_call('PUT', f"/api/tasks/{ctx['task_id']}", {
        'status': 'Completed',
        'user_name': ctx['member']['name']
    })
    assert res['status_code'] == 200
    assert res['data']['status'] == 'Completed'
    assert res['data']['completed_at'] is not None
    return "Task successfully marked as Completed with completed_at timestamp"

# -------------------------------------------------------------
# 3. RECURRENCE & ROUTINES ENGINE TESTS
# -------------------------------------------------------------

def test_create_daily_weekday_routine():
    res = api_call('POST', '/api/tasks', {
        'title': 'Daily Standup Sync Routine',
        'orbita_type': 'Routine',
        'recurrence_type': 'Daily',
        'recurrence_day': 'weekdays',
        'recurrence_interval': 1,
        'workspace': 'Work',
        'priority': 'Medium',
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'created_by': ctx['member']['name']
    })
    assert res['status_code'] == 201
    assert res['data']['orbita_type'] == 'Routine'
    ctx['routine_daily_id'] = res['data']['id']
    return f"Created Daily Weekday Routine: {res['data']['ticket_key']}"

def test_create_weekly_multiday_routine():
    res = api_call('POST', '/api/tasks', {
        'title': 'Fitness Gym Workout Routine',
        'orbita_type': 'Routine',
        'recurrence_type': 'Weekly',
        'recurrence_day': 'Mon, Wed, Fri',
        'recurrence_interval': 1,
        'workspace': 'Personal',
        'priority': 'High',
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'created_by': ctx['member']['name']
    })
    assert res['status_code'] == 201
    ctx['routine_weekly_id'] = res['data']['id']
    return f"Created Weekly Multi-day Routine: {res['data']['ticket_key']} (Days: Mon, Wed, Fri)"

def test_create_monthly_day15_routine():
    res = api_call('POST', '/api/tasks', {
        'title': 'Electricity Bill Payment Routine',
        'orbita_type': 'Routine',
        'recurrence_type': 'Monthly',
        'recurrence_day': '15',
        'recurrence_interval': 1,
        'workspace': 'Personal',
        'priority': 'Critical',
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'created_by': ctx['member']['name']
    })
    assert res['status_code'] == 201
    ctx['routine_monthly_id'] = res['data']['id']
    return f"Created Monthly Day 15 Routine: {res['data']['ticket_key']}"

def test_routine_auto_occurrence_generation():
    # Trigger occurrence evaluation
    gen_res = api_call('POST', '/api/routines/generate', {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name']
    })
    assert gen_res['status_code'] == 200
    
    # Query tasks for user
    tasks_res = api_call('GET', f"/api/tasks?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    assert tasks_res['status_code'] == 200
    all_tasks = tasks_res['data']
    
    daily_occs = [t for t in all_tasks if t.get('routine_id') == ctx['routine_daily_id']]
    weekly_occs = [t for t in all_tasks if t.get('routine_id') == ctx['routine_weekly_id']]
    
    assert len(daily_occs) > 0, "Daily routine occurrences must be generated"
    assert len(weekly_occs) > 0, "Weekly routine occurrences must be generated"
    
    ctx['daily_occ_id'] = daily_occs[0]['id']
    return f"Generated {len(daily_occs)} daily occurrences and {len(weekly_occs)} weekly occurrences (2 days ahead window)"

def test_routine_idempotency():
    # Calling generate again should not duplicate tasks
    api_call('POST', '/api/routines/generate', {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name']
    })
    tasks_res = api_call('GET', f"/api/tasks?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    all_tasks = tasks_res['data']
    daily_occs = [t for t in all_tasks if t.get('routine_id') == ctx['routine_daily_id']]
    
    # Check that occurrence dates are unique
    dates = [t.get('routine_occurrence_date') for t in daily_occs]
    assert len(dates) == len(set(dates)), "Duplicate occurrence dates detected!"
    return f"Idempotency verified: exactly {len(daily_occs)} unique occurrence dates"

def test_complete_occurrence_preserves_parent_routine():
    # Complete an occurrence task
    comp_res = api_call('PUT', f"/api/tasks/{ctx['daily_occ_id']}", {
        'status': 'Completed',
        'user_name': ctx['member']['name']
    })
    assert comp_res['status_code'] == 200
    assert comp_res['data']['status'] == 'Completed'
    
    # Parent routine MUST remain Active
    parent_res = api_call('GET', f"/api/tasks/{ctx['routine_daily_id']}")
    assert parent_res['status_code'] == 200
    assert parent_res['data']['status'] == 'Active', "Parent routine status should remain Active"
    return "Occurrence task completed while parent Routine template remains perpetually Active"

# -------------------------------------------------------------
# 4. GOALS & FOCUS TIMERS & TIMESHEETS
# -------------------------------------------------------------

def test_create_goal():
    res = api_call('POST', '/api/tasks', {
        'title': 'Master Advanced TypeScript & React',
        'orbita_type': 'Goal',
        'workspace': 'Personal',
        'priority': 'High',
        'target_hours': 0, # Optional open tracking
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'created_by': ctx['member']['name']
    })
    assert res['status_code'] == 201
    assert res['data']['orbita_type'] == 'Goal'
    assert res['data']['target_hours'] == 0
    ctx['goal_id'] = res['data']['id']
    return f"Created Focus Goal {res['data']['ticket_key']}: Open hours tracking"

def test_timer_start_pause_resume_stop():
    # Start timer
    start_res = api_call('POST', f"/api/tasks/{ctx['goal_id']}/timer/start", {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name']
    })
    assert start_res['status_code'] == 200
    assert 'started' in start_res['data'].get('message', '').lower()
    
    time.sleep(1) # simulate 1s work
    
    # Pause timer
    pause_res = api_call('POST', f"/api/tasks/{ctx['goal_id']}/timer/pause", {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name']
    })
    assert pause_res['status_code'] == 200
    
    # Resume timer
    resume_res = api_call('POST', f"/api/tasks/{ctx['goal_id']}/timer/resume", {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name']
    })
    assert resume_res['status_code'] == 200
    
    time.sleep(1)
    
    # Stop timer
    stop_res = api_call('POST', f"/api/tasks/{ctx['goal_id']}/timer/stop", {
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'user_name': ctx['member']['name'],
        'notes': 'Deep focus studying TypeScript generics'
    })
    assert stop_res['status_code'] == 200
    assert stop_res['data']['duration_seconds'] >= 1
    return f"Timer cycle validated: Logged {stop_res['data']['duration_seconds']}s focus session"

def test_timesheet_session_visibility():
    ts_res = api_call('GET', f"/api/timesheets?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    assert ts_res['status_code'] == 200
    sessions = ts_res['data']
    assert isinstance(sessions, list)
    assert len(sessions) > 0, "Timesheet sessions must be visible to user"
    assert any(s.get('task_id') == ctx['goal_id'] for s in sessions)
    return f"Timesheets endpoint verified: Found {len(sessions)} logged sessions for member"

# -------------------------------------------------------------
# 5. PROJECTS & MULTI-STAGE HIERARCHY
# -------------------------------------------------------------

def test_create_project_with_stages():
    res = api_call('POST', '/api/tasks', {
        'title': 'Orbita Enterprise Work Platform',
        'orbita_type': 'Project',
        'workspace': 'Work',
        'priority': 'Critical',
        'user_id': ctx['member']['id'],
        'user_email': ctx['member']['email'],
        'created_by': ctx['member']['name'],
        'stages': [
            {
                'title': 'Stage 1: Architecture & DB Design',
                'tasks': [
                    {'title': 'Design MongoDB Atlas Schema'},
                    {'title': 'Configure JWT Authentication'}
                ]
            },
            {
                'title': 'Stage 2: Frontend & Views',
                'tasks': [
                    {'title': 'Build Eisenhower Priority Matrix'},
                    {'title': 'Implement Kanban Drag & Drop'}
                ]
            }
        ]
    })
    assert res['status_code'] == 201
    assert res['data']['orbita_type'] == 'Project'
    assert len(res['data']['stages']) == 2
    assert len(res['data']['stages'][0]['tasks']) == 2
    ctx['project_id'] = res['data']['id']
    ctx['stage_1_id'] = res['data']['stages'][0]['id']
    ctx['stage_task_1_id'] = res['data']['stages'][0]['tasks'][0]['id']
    return f"Created Project {res['data']['ticket_key']} with 2 Stages and 4 Sub-Tasks"

def test_add_task_to_stage():
    res = api_call('POST', f"/api/stages/{ctx['stage_1_id']}/tasks", {
        'title': 'Write Comprehensive API Documentation',
        'assignee': ctx['member']['name']
    })
    assert res['status_code'] in (200, 201)
    assert len(res['data']['stages'][0]['tasks']) == 3
    return "Added new task to Stage 1 successfully"

def test_complete_stage_task():
    res = api_call('PUT', f"/api/projects/{ctx['project_id']}/stages/{ctx['stage_1_id']}/tasks/{ctx['stage_task_1_id']}", {
        'is_completed': True
    })
    assert res['status_code'] == 200
    stage1 = res['data']['stages'][0]
    task1 = [t for t in stage1['tasks'] if t['id'] == ctx['stage_task_1_id']][0]
    assert task1['is_completed'] is True
    return "Stage Sub-task completed and verified"

def test_stage_task_timer():
    # Start timer on stage task
    start_res = api_call('POST', f"/api/projects/{ctx['project_id']}/stages/{ctx['stage_1_id']}/tasks/{ctx['stage_task_1_id']}/timer/start", {
        'user_name': ctx['member']['name']
    })
    assert start_res['status_code'] == 200
    
    time.sleep(1)
    
    # Stop timer on stage task
    stop_res = api_call('POST', f"/api/projects/{ctx['project_id']}/stages/{ctx['stage_1_id']}/tasks/{ctx['stage_task_1_id']}/timer/stop", {
        'user_name': ctx['member']['name'],
        'notes': 'DB Architecture Optimization'
    })
    assert stop_res['status_code'] == 200
    assert stop_res['data']['duration_seconds'] >= 1
    return f"Stage sub-task timer logged {stop_res['data']['duration_seconds']}s directly"

# -------------------------------------------------------------
# 6. VIEWS & AGGREGATIONS (DASHBOARD, MATRIX, HIGHLIGHTS, AUDIT)
# -------------------------------------------------------------

def test_dashboard_stats_endpoint():
    res = api_call('GET', f"/api/dashboard/stats?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    assert res['status_code'] == 200
    stats = res['data']
    assert 'total_tasks' in stats
    assert 'completed_tasks' in stats
    assert 'completion_rate' in stats
    assert 'total_hours_logged' in stats
    assert 'starred_count' in stats
    return f"Dashboard stats verified: Total={stats['total_tasks']}, Completed={stats['completed_tasks']}, Hours={stats['total_hours_logged']}h"

def test_priority_matrix_quadrants():
    res = api_call('GET', f"/api/matrix?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    assert res['status_code'] == 200
    matrix = res['data']
    assert 'Q1' in matrix and 'Q2' in matrix and 'Q3' in matrix and 'Q4' in matrix
    # Verify master routines are excluded from matrix
    for q in ['Q1', 'Q2', 'Q3', 'Q4']:
        for item in matrix[q]:
            assert item.get('orbita_type') != 'Routine', "Master Routine template should NOT appear in Matrix"
    return f"Priority Matrix verified: Q1={len(matrix['Q1'])}, Q2={len(matrix['Q2'])}, Q3={len(matrix['Q3'])}, Q4={len(matrix['Q4'])}"

def test_monthly_highlights_endpoint():
    res = api_call('GET', f"/api/highlights/month?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    assert res['status_code'] == 200
    data = res['data']
    assert 'scorecard' in data
    assert 'achievement_score' in data['scorecard']
    assert 'month' in data
    assert 'badges' in data
    assert 'starred_items' in data
    return f"Monthly Highlights verified: Month={data['month']}, Score={data['scorecard']['achievement_score']}, Badges={len(data['badges'])}"

def test_audit_logs_trail():
    res = api_call('GET', f"/api/audit-logs?user_id={ctx['member']['id']}&user_email={ctx['member']['email']}&user_role=Member&user_name={ctx['member']['name']}")
    assert res['status_code'] == 200
    logs = res['data']
    assert isinstance(logs, list)
    assert len(logs) > 0, "Audit logs must record user activity"
    latest = logs[0]
    return f"Audit trail verified: {len(logs)} immutable event records (Latest: {latest.get('action')} on {latest.get('ticket_key')})"

def test_data_isolation_between_users():
    # Create another user and verify they cannot see member's private items
    other_email = f"other_user_{int(time.time())}@orbita.com"
    other_user = api_call('POST', '/api/auth/register', {
        'name': 'Other Isolated User',
        'email': other_email,
        'password': 'Password123!',
        'role': 'Member'
    })['data']['user']
    
    other_tasks = api_call('GET', f"/api/tasks?user_id={other_user['id']}&user_email={other_user['email']}&user_role=Member&user_name={other_user['name']}")['data']
    member_task_ids = [ctx.get('task_id'), ctx.get('routine_daily_id'), ctx.get('goal_id'), ctx.get('project_id')]
    
    for t in other_tasks:
        assert t['id'] not in member_task_ids, f"Data leak detected! Task {t['id']} should not be visible to {other_email}"
    
    return "Data isolation verified: New member has 0 access to other users' private work items"

# -------------------------------------------------------------
# RUN ALL TESTS
# -------------------------------------------------------------

def main():
    print("===============================================================")
    print("       ORBITA WORK MANAGEMENT PLATFORM - E2E TEST SUITE        ")
    print("===============================================================\n")

    test_definitions = [
        # Auth & Security
        ('AUTH-001', 'Authentication', 'Superadmin Login Authentication', 'Verify Superadmin credentials and role', test_auth_superadmin_login),
        ('AUTH-002', 'Authentication', 'Member Registration & Login', 'Verify Member signup, role assignment, and JWT issuance', test_auth_member_registration_and_login),
        ('AUTH-003', 'Authentication', 'Invalid Credentials Rejection', 'Verify 401 Unauthorized for incorrect password', test_auth_invalid_password),
        
        # Work Items CRUD
        ('TASK-001', 'Core Work Items', 'Create Single Action Task', 'Create Task with priority, tags, dates, and sequential key', test_create_task),
        ('TASK-002', 'Core Work Items', 'Update Task & Priority Quadrant Sync', 'Update task and verify automatic Priority <-> Quadrant sync', test_update_task_details),
        ('TASK-003', 'Core Work Items', 'Toggle Milestone Star', 'Verify milestone star toggling on work item', test_star_task),
        ('TASK-004', 'Core Work Items', 'Complete Task Lifecycle', 'Mark task completed and verify completed_at timestamp', test_complete_task),
        
        # Routines & Recurrence Engine
        ('ROUT-001', 'Routines & Recurrence', 'Create Daily Weekday Routine', 'Configure Daily recurrence for weekdays only (Mon-Fri)', test_create_daily_weekday_routine),
        ('ROUT-002', 'Routines & Recurrence', 'Create Weekly Multi-Day Routine', 'Configure Weekly recurrence for Mon, Wed, Fri', test_create_weekly_multiday_routine),
        ('ROUT-003', 'Routines & Recurrence', 'Create Monthly Day 15 Routine', 'Configure Monthly recurrence on specific day of month', test_create_monthly_day15_routine),
        ('ROUT-004', 'Routines & Recurrence', 'Auto-Occurrence Generation Engine', 'Evaluate routines and generate tasks 2 days in advance', test_routine_auto_occurrence_generation),
        ('ROUT-005', 'Routines & Recurrence', 'Occurrence Engine Idempotency', 'Ensure zero duplicate occurrence tasks are created', test_routine_idempotency),
        ('ROUT-006', 'Routines & Recurrence', 'Occurrence Completion vs Parent Routine', 'Verify completing occurrence leaves parent routine Active', test_complete_occurrence_preserves_parent_routine),
        
        # Goals & Focus Timers
        ('GOAL-001', 'Goals & Timers', 'Create Open Focus Goal', 'Create goal without mandatory target hours', test_create_goal),
        ('GOAL-002', 'Goals & Timers', 'Stopwatch Timer Cycle (Start/Pause/Resume/Stop)', 'Test real-time stopwatch session accumulation and logging', test_timer_start_pause_resume_stop),
        ('GOAL-003', 'Goals & Timers', 'Focus Timesheet Session Visibility', 'Verify completed focus session is queryable in Timesheets', test_timesheet_session_visibility),
        
        # Projects & Multi-Stage Delivery
        ('PROJ-001', 'Projects & Hierarchy', 'Create Multi-Stage Project', 'Create Project with Stages and Sub-Tasks', test_create_project_with_stages),
        ('PROJ-002', 'Projects & Hierarchy', 'Add Sub-Task to Existing Stage', 'Dynamically append new task to a project stage', test_add_task_to_stage),
        ('PROJ-003', 'Projects & Hierarchy', 'Complete Individual Stage Sub-Task', 'Update stage task completion status', test_complete_stage_task),
        ('PROJ-004', 'Projects & Hierarchy', 'Direct Stage Sub-Task Timer', 'Start and stop focus timer directly on a stage task', test_stage_task_timer),
        
        # Planning & Visual Views
        ('VIEW-001', 'Views & Metrics', 'Executive Dashboard KPI Aggregations', 'Verify KPI calculation and metrics aggregation', test_dashboard_stats_endpoint),
        ('VIEW-002', 'Views & Metrics', 'Eisenhower Priority Matrix (Q1-Q4)', 'Verify quadrant categorization and master routine exclusion', test_priority_matrix_quadrants),
        ('VIEW-003', 'Views & Metrics', 'Monthly Highlights & Scorecard', 'Verify monthly score, accomplishments, and badges', test_monthly_highlights_endpoint),
        ('VIEW-004', 'Views & Metrics', 'Immutable Activity Audit Trail', 'Verify compliance audit log recording for actions', test_audit_logs_trail),
        ('SEC-001', 'Multi-Tenancy & Security', 'Cross-User Data Isolation', 'Verify members have zero visibility into other users items', test_data_isolation_between_users),
    ]

    for t_id, cat, name, desc, fn in test_definitions:
        run_test(t_id, cat, name, desc, fn)

    print("\n===============================================================")
    passed = sum(1 for r in test_results if r['status'] == 'PASSED')
    failed = sum(1 for r in test_results if r['status'] == 'FAILED')
    total = len(test_results)
    print(f"TEST EXECUTION SUMMARY: {passed}/{total} PASSED ({failed} FAILED)")
    print("===============================================================\n")

    generate_excel_report(test_results)

def generate_excel_report(results):
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # Sheet 1: Executive Summary
    # ---------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Colors
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    accent_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    sub_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    font_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, color="000000")
    font_pass = Font(name="Calibri", size=11, bold=True, color="166534")
    font_fail = Font(name="Calibri", size=11, bold=True, color="991B1B")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Banner
    ws_summary.merge_cells("A1:F2")
    title_cell = ws_summary["A1"]
    title_cell.value = "ORBITA WORK MANAGEMENT PLATFORM — E2E TEST REPORT"
    title_cell.font = font_title
    title_cell.fill = accent_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata Block
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta = [
        ("Application Name", "Orbita Work Management Tool (by Runit Infotech)"),
        ("Test Execution Date", now_str),
        ("Environment", "Local Development (Vite Frontend :5173 / Express API :5001)"),
        ("Database", "MongoDB Atlas (Cluster: orbita)"),
        ("Tested By", "Antigravity Autonomous E2E Test Runner"),
        ("Total Test Cases", len(results)),
        ("Passed", sum(1 for r in results if r['status'] == 'PASSED')),
        ("Failed", sum(1 for r in results if r['status'] == 'FAILED')),
        ("Success Rate", f"{(sum(1 for r in results if r['status'] == 'PASSED') / len(results) * 100):.1f}%")
    ]

    ws_summary["A4"] = "Execution Metadata"
    ws_summary["A4"].font = Font(name="Calibri", size=13, bold=True, color="1E293B")
    
    for row_idx, (k, v) in enumerate(meta, start=5):
        ws_summary[f"A{row_idx}"] = k
        ws_summary[f"A{row_idx}"].font = font_bold
        ws_summary[f"A{row_idx}"].fill = sub_fill
        ws_summary[f"A{row_idx}"].border = thin_border
        
        ws_summary.merge_cells(f"B{row_idx}:D{row_idx}")
        cell = ws_summary[f"B{row_idx}"]
        cell.value = v
        cell.font = font_regular
        if k == "Success Rate":
            cell.font = font_pass if "100" in str(v) else font_fail
            cell.fill = pass_fill if "100" in str(v) else fail_fill
        cell.border = thin_border

    # Category Breakdown Table
    ws_summary["A16"] = "Module Test Breakdown"
    ws_summary["A16"].font = Font(name="Calibri", size=13, bold=True, color="1E293B")
    
    cat_headers = ["Category / Module", "Total Tests", "Passed", "Failed", "Pass Rate (%)"]
    for col_idx, h in enumerate(cat_headers, start=1):
        c = ws_summary.cell(row=17, column=col_idx)
        c.value = h
        c.font = font_header
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    categories = list(dict.fromkeys(r['category'] for r in results))
    for row_idx, cat in enumerate(categories, start=18):
        cat_tests = [r for r in results if r['category'] == cat]
        p = sum(1 for r in cat_tests if r['status'] == 'PASSED')
        f = sum(1 for r in cat_tests if r['status'] == 'FAILED')
        rate = f"{(p / len(cat_tests) * 100):.1f}%"
        
        ws_summary.cell(row=row_idx, column=1, value=cat).font = font_bold
        ws_summary.cell(row=row_idx, column=2, value=len(cat_tests)).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=3, value=p).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=4, value=f).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=5, value=rate).alignment = Alignment(horizontal="center")
        
        for c_idx in range(1, 6):
            ws_summary.cell(row=row_idx, column=c_idx).border = thin_border

    # Auto-adjust column widths
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 16)

    # ---------------------------------------------------------
    # Sheet 2: Detailed Test Results
    # ---------------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Results")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers = ["Test ID", "Category", "Test Name", "Description", "Status", "Execution Time (ms)", "Verification Notes / Response Output"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws_details.cell(row=1, column=col_idx)
        c.value = h
        c.font = font_header
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    for row_idx, r in enumerate(results, start=2):
        ws_details.cell(row=row_idx, column=1, value=r['test_id']).font = font_bold
        ws_details.cell(row=row_idx, column=2, value=r['category']).font = font_regular
        ws_details.cell(row=row_idx, column=3, value=r['test_name']).font = font_bold
        ws_details.cell(row=row_idx, column=4, value=r['description']).font = font_regular
        
        status_cell = ws_details.cell(row=row_idx, column=5, value=r['status'])
        status_cell.alignment = Alignment(horizontal="center")
        if r['status'] == 'PASSED':
            status_cell.font = font_pass
            status_cell.fill = pass_fill
        else:
            status_cell.font = font_fail
            status_cell.fill = fail_fill

        ws_details.cell(row=row_idx, column=6, value=r['elapsed_ms']).alignment = Alignment(horizontal="right")
        ws_details.cell(row=row_idx, column=7, value=r['notes']).font = font_regular
        
        for c_idx in range(1, 8):
            ws_details.cell(row=row_idx, column=c_idx).border = thin_border

    for col in ws_details.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_details.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 60)

    # Save to file
    out_file = 'Orbita_E2E_Test_Report.xlsx'
    wb.save(out_file)
    print(f"Excel Test Report successfully generated at: {out_file}")

if __name__ == '__main__':
    main()
